import gurobipy as gp
import numpy
from gurobipy import *
import numpy as np
import pandas as pd
import time
import pwlf
from pwlf import *
import pickle  # 用于存储变量F和T
import os
import matplotlib.pyplot as plt
from openpyxl import Workbook
from pypower import case118
plt.rcParams['font.sans-serif'] = ['SimHei']  # 用于正常显示中文标签
plt.rcParams['axes.unicode_minus'] = False  # 用于正常显示负号

class Cases:
    def __init__(self):
        self.Info = None
        self.Dt = None
        self.sigma = None

    def read_col(self, recarray, col_name):
        col = recarray[col_name]
        return np.squeeze(col.reshape(len(col),1))

    def read_data(self, file_path, sheet_name):
        with pd.ExcelFile(file_path) as xsl:
            df = pd.read_excel(file_path, sheet_name)
        self.Info = df.to_records(index=False)

        self.Dt = self.read_col(self.Info, '预计负荷')
        self.sigma = self.read_col(self.Info, '负载可能变化')

class Line:
    def __init__(self):
        self.Info = None
        self.Rdc = None
        self.TM = None
        self.X = None
        self.Q = None
        self.Pdcmin = None
        self.Pdcmax = None
        self.delta = None
        self.cw = None

    def read_col(self, recarray, col_name):
        col = recarray[col_name]
        return np.squeeze(col.reshape(len(col), 1))

    def read_data(self, file_path, sheet_name):
        with pd.ExcelFile(file_path) as xsl:
            df = pd.read_excel(file_path, sheet_name)
        self.Info = df.to_records(index=False)

        self.Rdc = self.read_col(self.Info, '爬坡限制')
        self.TM = self.read_col(self.Info, '调整最小持续时间')
        self.X = self.read_col(self.Info, '最大调整次数')
        self.Q = self.read_col(self.Info, '直流联络日交易电量')
        self.Pdcmin = self.read_col(self.Info, '传输最小功率')
        self.Pdcmax = self.read_col(self.Info, '传输最大功率')
        self.delta = self.read_col(self.Info, '模拟时间步长')
        self.cw = self.read_col(self.Info, '弃风惩罚系数')

class SU:
    def __init__(self):
        self.Info = None
        self.Pmin = None
        self.Pmax = None
        self.c1 = None
        self.c2 = None
        self.c3 = None
        self.tv1 = None
        self.tv2 = None
        self.NL = None
        self.Cop = None
        self.Csd = None
        self.Units = None
        self.UT = None
        self.DT = None
        self.RU = None
        self.v0 = None
        self.V0 = None
        self.U0 = None
        self.S0 = None

    def read_col(self, recarray, col_name):
        col = recarray[col_name]
        return np.squeeze(col.reshape(len(col), 1))

    def read_data(self, file_path, sheet_name):
        with pd.ExcelFile(file_path) as xsl:
            df = pd.read_excel(file_path, sheet_name)
        self.Info = df.to_records(index=False)

        self.Pmin = self.read_col(self.Info, '最小出力')
        self.Pmax = self.read_col(self.Info, '最大出力')
        self.c1 = self.read_col(self.Info, '燃料费用系数1')
        self.c2 = self.read_col(self.Info, '燃料费用系数2')
        self.c3 = self.read_col(self.Info, '燃料费用系数3')
        self.tv1 = self.read_col(self.Info, '出力阈值1')
        self.tv2 = self.read_col(self.Info, '出力阈值2')
        self.NL = self.read_col(self.Info, '燃料费用分段数')
        self.Cop = self.read_col(self.Info, '开机成本')
        self.Csd = self.read_col(self.Info, '关机成本')
        self.Units = self.read_col(self.Info, '机组编号') - 1
        self.UT = self.read_col(self.Info, '最小开机时间')
        self.DT = self.read_col(self.Info, '最小关机时间')
        self.RU = self.read_col(self.Info, '爬坡能力')
        self.v0 = self.read_col(self.Info, '初始状态')
        self.V0 = (self.v0 > 0).astype(int)
        self.U0 = np.where(self.v0 < 0, 0, self.v0)
        self.S0 = np.where(self.v0 > 0, 0, -self.v0)

class RU:
    def __init__(self):
        self.Info = None
        self.Pmin = None
        self.Pmax = None
        self.c1 = None
        self.c2 = None
        self.c3 = None
        self.tv1 = None
        self.tv2 = None
        self.NL = None
        self.Cop = None
        self.Csd = None
        self.Units = None
        self.UT = None
        self.DT = None
        self.RU = None
        self.v0 = None
        self.V0 = None
        self.U0 = None
        self.S0 = None

    def read_col(self, recarray, col_name):
        col = recarray[col_name]
        return np.squeeze(col.reshape(len(col), 1))

    def read_data(self, file_path, sheet_name):
        with pd.ExcelFile(file_path) as xsl:
            df = pd.read_excel(file_path, sheet_name)
        self.Info = df.to_records(index=False)

        self.Pmin = self.read_col(self.Info, '最小出力')
        self.Pmax = self.read_col(self.Info, '最大出力')
        self.c1 = self.read_col(self.Info, '燃料费用系数1')
        self.c2 = self.read_col(self.Info, '燃料费用系数2')
        self.c3 = self.read_col(self.Info, '燃料费用系数3')
        self.tv1 = self.read_col(self.Info, '出力阈值1')
        self.tv2 = self.read_col(self.Info, '出力阈值2')
        self.NL = self.read_col(self.Info, '燃料费用分段数')
        self.Cop = self.read_col(self.Info, '开机成本')
        self.Csd = self.read_col(self.Info, '关机成本')
        self.Units = self.read_col(self.Info, '机组编号') - 1
        self.UT = self.read_col(self.Info, '最小开机时间')
        self.DT = self.read_col(self.Info, '最小关机时间')
        self.RU = self.read_col(self.Info, '爬坡能力')
        self.v0 = self.read_col(self.Info, '初始状态')
        self.V0 = (self.v0 > 0).astype(int)
        self.U0 = np.where(self.v0 < 0, 0, self.v0)
        self.S0 = np.where(self.v0 > 0, 0, -self.v0)

class SNE:
    def __init__(self):
        self.Info = None
        self.Pwpred = None

    def read_col(self, recarray, col_name):
        col = recarray[col_name]
        return np.squeeze(col.reshape(len(col), 1))

    def read_data(self, file_path, sheet_name):
        with pd.ExcelFile(file_path) as xsl:
            df = pd.read_excel(file_path, sheet_name)
        self.Info = df.to_records(index=False)

        self.Pwpred = self.read_col(self.Info, '预计出力')

def read(file_path):
    cases.read_data(file_path, 'Case')
    line.read_data(file_path, 'Line')
    su.read_data(file_path, 'Sending_Units')
    ru.read_data(file_path, 'Receiving_Units')
    sne.read_data(file_path, 'Sending_NE')

def set_params():

    t = len(cases.Info)
    j = len(su.Info)
    k = len(ru.Info)

    params['time_interval'] = t

    params['T'] = range(t)
    params['La'] = range(su.NL[0])
    params['Lb'] = range(ru.NL[0])
    params['Na'] = range(su.NL[0]-1)
    params['Nb'] = range(ru.NL[0]-1)
    params['su_unit_num'] = range(j)
    params['ru_unit_num'] = range(k)
    params['Su_unit_num'] = j
    params['Ru_unit_num'] = k

MP = gp.Model(name = 'MP')
cases = Cases()
line = Line()
su = SU()D:\
ru = RU()
sne = SNE()
read(r".\Data\P10Data.xlsx")
params = {}
set_params()

eta = MP.addVar(lb=0,vtype=GRB.CONTINUOUS,name='eta')
cau = MP.addVars(params['Su_unit_num'], params['time_interval'], lb=0, vtype=GRB.CONTINUOUS, name='cau')
cbu = MP.addVars(params['Ru_unit_num'], params['time_interval'], lb=0, vtype=GRB.CONTINUOUS, name='cbu')
cad = MP.addVars(params['Su_unit_num'], params['time_interval'], lb=0, vtype=GRB.CONTINUOUS, name='cad')
cbd = MP.addVars(params['Ru_unit_num'], params['time_interval'], lb=0, vtype=GRB.CONTINUOUS, name='cbd')

xa = MP.addVars(params['Su_unit_num'], params['time_interval'], vtype=GRB.BINARY, name='xa')
xb = MP.addVars(params['Ru_unit_num'], params['time_interval'], vtype=GRB.BINARY, name='xb')
vdc = MP.addVars(params['time_interval'], vtype=GRB.BINARY, name='vdc')
MP.update()

y = {}
Xa,Xb,Vdc = {},{},{}
u = {}
var_dict_MP = {}

def add_MP_var(crim):

    var_dict_MP[crim,'ca'] = MP.addVars(params['time_interval'],lb=0, vtype=GRB.CONTINUOUS)
    var_dict_MP[crim,'cb'] = MP.addVars(params['time_interval'],lb=0, vtype=GRB.CONTINUOUS)
    var_dict_MP[crim,'deltaa'] = MP.addVars(params['Su_unit_num'], params['La'], params['time_interval'],lb=0, vtype=GRB.CONTINUOUS)
    var_dict_MP[crim,'deltab'] = MP.addVars(params['Ru_unit_num'], params['Lb'], params['time_interval'],lb=0, vtype=GRB.CONTINUOUS)
    var_dict_MP[crim,'PA'] = MP.addVars(params['Su_unit_num'], params['time_interval'],lb=0, vtype=GRB.CONTINUOUS)
    var_dict_MP[crim,'PB'] = MP.addVars(params['Ru_unit_num'], params['time_interval'],lb=0, vtype=GRB.CONTINUOUS)
    var_dict_MP[crim,'PAMAX'] = MP.addVars(params['Su_unit_num'], params['time_interval'],lb=0, vtype=GRB.CONTINUOUS)
    var_dict_MP[crim,'PBMAX'] = MP.addVars(params['Ru_unit_num'], params['time_interval'],lb=0, vtype=GRB.CONTINUOUS)
    var_dict_MP[crim,'Pwact'] = MP.addVars(params['time_interval'],lb=0, vtype=GRB.CONTINUOUS)
    var_dict_MP[crim,'Pwcur'] = MP.addVars(params['time_interval'],lb=0, vtype=GRB.CONTINUOUS)
    var_dict_MP[crim,'Pdc'] = MP.addVars(params['time_interval'],lb=0, vtype=GRB.CONTINUOUS)
    var_dict_MP[crim,'cw'] = MP.addVars(params['time_interval'],lb=0, vtype=GRB.CONTINUOUS)
    MP.update()





def add_MP_cost_constr(var,crim):

    PA = var[crim,'PA']
    PB = var[crim,'PB']

    ca = var[crim,'ca']
    cb = var[crim,'cb']
    deltaa = var[crim,'deltaa']
    deltab = var[crim,'deltab']



    cw = var[crim,'cw']


    Pwcur = var[crim,'Pwcur']

    T = params['T']
    La = params['La']
    Lb = params['Lb']

    J = params['su_unit_num']
    K = params['ru_unit_num']

    suc1 = su.c1
    suc2 = su.c2
    suc3 = su.c3
    sutv1 = su.tv1
    sutv2 = su.tv2
    ruc1 = ru.c1
    ruc2 = ru.c2
    ruc3 = ru.c3
    rutv1 = ru.tv1
    rutv2 = ru.tv2
    suNL = len(La)
    ruNL = len(Lb)
    suop = su.Cop
    susd = su.Csd
    ruop = ru.Cop
    rusd = ru.Csd
    suV0 = su.V0
    ruV0 = ru.V0
    Pamax = su.Pmax
    Pbmax = ru.Pmax

    FA = np.column_stack([suc1, suc2, suc3])
    FB = np.column_stack([ruc1, ruc2, ruc3])
    TVA = np.column_stack([sutv1, sutv2])
    TVB = np.column_stack([rutv1, rutv2])

    MP.addConstrs(ca[t] == gp.quicksum(FA[j,l] * deltaa[j,l,t] for j in J for l in La) for t in T)
    MP.addConstrs(PA[j,t] == gp.quicksum(deltaa[j,l,t] for l in La) for j in J for t in T)
    MP.addConstrs(deltaa[j,0,t] <= TVA[j,0] for t in T for j in J)
    MP.addConstrs(deltaa[j,suNL-1,t] <= Pamax[j] - TVA[j,suNL-2] for t in T for j in J)
    MP.addConstrs(deltaa[j,l,t] <= TVA[j,l] - TVA[j,l-1] for j in J for t in T for l in La[1:suNL-1])


    MP.addConstrs(cb[t] == gp.quicksum(FB[k, l] * deltab[k, l, t] for k in K for l in Lb) for t in T)
    MP.addConstrs(PB[k, t] == gp.quicksum(deltab[k, l, t] for l in Lb) for k in K for t in T)
    MP.addConstrs(deltab[k, 0, t] <= TVB[k, 0] for t in T for k in K)
    MP.addConstrs(deltab[k, ruNL - 1, t] <= Pbmax[k] - TVB[k, ruNL - 2] for t in T for k in K)
    MP.addConstrs(deltab[k, l, t] <= TVB[k, l] - TVB[k, l - 1] for k in K for t in T for l in Lb[1:suNL - 1])





    MP.addConstrs(cw[t] == line.cw * Pwcur[t] for t in T)



def add_MP_RUN_constr(var,crim):
    PA = var[crim,'PA']
    PB = var[crim,'PB']
    PAMAX = var[crim,'PAMAX']
    PBMAX = var[crim,'PBMAX']


    suPmax = su.Pmax
    suPmin = su.Pmin
    ruPmax = su.Pmax
    ruPmin = ru.Pmin
    suRU = su.RU
    ruRU = ru.RU
    suUT = su.UT
    ruUT = ru.UT
    suDT = su.DT
    ruDT = ru.DT
    suU0 = su.U0
    suS0 = su.S0
    ruU0 = ru.U0
    ruS0 = ru.S0
    suV0 = su.V0
    ruV0 = ru.V0
    Pamax = su.Pmax
    Pbmax = ru.Pmax

    T = params['T']
    J = params['su_unit_num']
    K = params['ru_unit_num']
    Tia = np.squeeze(np.full((1, len(J)), params['time_interval']))
    Tib = np.squeeze(np.full((1, len(K)), params['time_interval']))
    GA = np.minimum(Tia, (suUT - suU0) * suV0)
    GB = np.minimum(Tib, (ruUT - ruU0) * ruV0)
    GA[GA<0] =0
    GB[GB<0]=0
    LA = np.minimum(Tia, (suDT - suS0) * (1 - suV0))
    LB = np.minimum(Tib, (ruDT - ruS0) * (1 - ruV0))
    LA[LA<0] = 0
    LB[LB<0] = 0

    MP.addConstrs(PA[j, t] <= PAMAX[j, t] for t in T for j in J)
    MP.addConstrs(PA[j, t] >= suPmin[j] * xa[j, t] for t in T for j in J)
    MP.addConstrs(PB[k, t] <= PBMAX[k, t] for t in T for k in K)
    MP.addConstrs(PB[k, t] >= ruPmin[k] * xb[k, t] for k in K for t in T)
    MP.addConstrs(PAMAX[j, t] <= Pamax[j] * xa[j, t] for j in J for t in T)
    MP.addConstrs(PBMAX[k, t] <= Pbmax[k] * xb[k, t] for k in K for t in T)
    MP.addConstrs(
        PAMAX[j, t] <= PA[j, t-1] + suRU[j] * xa[j, t - 1] + suPmin[j] * (xa[j, t] - xa[j, t - 1]) + suPmax[j] * (
                    1 - xa[j, t]) for j in J for t in T[1:])
    MP.addConstrs(
        PBMAX[k, t] <= PB[k, t-1] + ruRU[k] * xb[k, t - 1] + ruPmin[k] * (xb[k, t] - xb[k, t - 1]) + ruPmax[k] * (
                    1 - xb[k, t]) for k in K for t in T[1:])


def add_MP_HVDC_constr(var,crim):
    Pdc = var[crim,'Pdc']

    Rdc = line.Rdc
    Pdcmin = line.Pdcmin
    Pdcmax = line.Pdcmax
    TM = line.TM
    X = line.X
    Q = line.Q
    delta = line.delta
    T = params['T']
    time_interval = params['time_interval']
    MP.addConstrs(Pdc[t] - Pdc[t-1] >= -vdc[t]*Rdc for t in T[1:])
    MP.addConstrs(Pdc[t] - Pdc[t - 1] <= vdc[t] * Rdc for t in T[1:])
    MP.addConstrs(Pdc[t] >= Pdcmin for t in T)
    MP.addConstrs(Pdc[t] <= Pdcmax for t in T)

    MP.addConstr(gp.quicksum(Pdc[t] * delta for t in T) == Q)

def add_MP_SEC_constr(var,crim):
    PA = var[crim,'PA']
    Pwact = var[crim,'Pwact']
    Pdc = var[crim,'Pdc']
    Pwcur = var[crim,'Pwcur']
    Pwpred = sne.Pwpred

    T = params['T']
    J = params['su_unit_num']
    MP.addConstrs(gp.quicksum(PA[j,t] for j in J) + Pwact[t] == Pdc[t] for t in T)
    MP.addConstrs(Pwact[t] <= Pwpred[t] for t in T)
    MP.addConstrs(Pwcur[t] == Pwpred[t] - Pwact[t] for t in T)

def add_MP_REC_constr(var,u,crim):
    PB = var[crim,'PB']
    Pdc = var[crim,'Pdc']

    Dt = cases.Dt
    sigma = cases.sigma

    T = params['T']
    K = params['ru_unit_num']
    MP.addConstrs(gp.quicksum(PB[k,t] for k in K) + Pdc[t] == Dt[t] + u[crim,t] * sigma[t] for t in T)



def add_MP_constr(var,u,crim):
    add_MP_cost_constr(var,crim)
    add_MP_RUN_constr(var,crim)
    add_MP_HVDC_constr(var,crim)
    add_MP_SEC_constr(var,crim)
    add_MP_REC_constr(var, u,crim)
def MP_solve(u,crim):
    add_MP_var(crim)

    ca = var_dict_MP[crim,'ca']
    cb = var_dict_MP[crim,'cb']
    cw = var_dict_MP[crim,'cw']
    T = params['T']
    J = params['su_unit_num']
    K = params['ru_unit_num']
    MP.setObjective(gp.quicksum(cau[j,t] + cad[j,t] for j in J for t in T)+ gp.quicksum(cbu[k,t]+ cbd[k,t] for k in K for t in T) +eta, gurobipy.GRB.MINIMIZE)
    add_MP_constr(var_dict_MP,u,crim)
    MP.addConstr(eta >= gp.quicksum(ca[t] + cb[t] + cw[t] for t in T))



    MP.optimize()


SP = gp.Model(name='SP')
M = 10000000000

def add_SP_var():
    global var_dict_SP
    var_dict_SP = {}
    var_dict_SP['Ca'] = SP.addVar(vtype=GRB.CONTINUOUS,lb=0, name='Ca')
    var_dict_SP['Cb'] = SP.addVar(vtype=GRB.CONTINUOUS,lb=0, name='Cb')
    var_dict_SP['ca'] = SP.addVars(params['time_interval'],lb=0,  vtype=GRB.CONTINUOUS, name='ca')
    var_dict_SP['cb'] = SP.addVars(params['time_interval'],lb=0, vtype=GRB.CONTINUOUS, name='cb')
    var_dict_SP['deltaa'] = SP.addVars(params['Su_unit_num'], params['La'], params['time_interval'],lb=0, vtype=GRB.CONTINUOUS, name='deltaa')
    var_dict_SP['deltab'] = SP.addVars(params['Ru_unit_num'], params['Lb'], params['time_interval'],lb=0,
                                       vtype=GRB.CONTINUOUS, name='deltab')

    var_dict_SP['PA'] = SP.addVars(params['Su_unit_num'], params['time_interval'],lb=0,  vtype=GRB.CONTINUOUS, name='PA')
    var_dict_SP['PB'] = SP.addVars(params['Ru_unit_num'], params['time_interval'],lb=0,  vtype=GRB.CONTINUOUS, name='PB')
    var_dict_SP['PAMAX'] = SP.addVars(params['Su_unit_num'], params['time_interval'],lb=0,  vtype=GRB.CONTINUOUS, name='PAMAX')
    var_dict_SP['PBMAX'] = SP.addVars(params['Ru_unit_num'], params['time_interval'],lb=0,  vtype=GRB.CONTINUOUS, name='PBMAX')
    var_dict_SP['Pwact'] = SP.addVars(params['time_interval'],lb=0,  vtype=GRB.CONTINUOUS, name='Pwact')
    var_dict_SP['Pwcur'] = SP.addVars(params['time_interval'],lb=0,  vtype=GRB.CONTINUOUS, name='Pwcur')
    var_dict_SP['Pdc'] = SP.addVars(params['time_interval'],lb=0, vtype=GRB.CONTINUOUS, name='Pdc')
    var_dict_SP['cw'] = SP.addVars(params['time_interval'],lb=0,  vtype=GRB.CONTINUOUS, name='cw')
    var_dict_SP['Cw'] = SP.addVar(vtype=GRB.CONTINUOUS,lb=0,  name='Cw')
    var_dict_SP['u'] = SP.addVars(params['time_interval'],lb=-1, ub=1,  vtype=GRB.CONTINUOUS, name='u')
    var_dict_SP['omega1'] = SP.addVars(params['Su_unit_num'], params['time_interval'], lb = 0, vtype=GRB.CONTINUOUS, name='omega1')
    var_dict_SP['omega2'] = SP.addVars(params['Su_unit_num'], params['time_interval'], lb=0,
                                       vtype=GRB.CONTINUOUS, name='omega2')
    var_dict_SP['omega3'] = SP.addVars(params['Su_unit_num'], params['La'], params['time_interval'], lb=0,
                                       vtype=GRB.CONTINUOUS, name='omega3')
    var_dict_SP['omega4'] = SP.addVars(params['Ru_unit_num'], params['time_interval'], lb=0,
                                       vtype=GRB.CONTINUOUS, name='omega4')
    var_dict_SP['omega5'] = SP.addVars(params['Ru_unit_num'], params['time_interval'], lb=0,
                                       vtype=GRB.CONTINUOUS, name='omega5')
    var_dict_SP['omega6'] = SP.addVars(params['Ru_unit_num'], params['Lb'], params['time_interval'], lb=0,
                                       vtype=GRB.CONTINUOUS, name='omega6')
    var_dict_SP['omega7'] = SP.addVars(params['Su_unit_num'], params['time_interval'], lb=0,
                                       vtype=GRB.CONTINUOUS, name='omega7')
    var_dict_SP['omega8'] = SP.addVars(params['Su_unit_num'], params['time_interval'], lb=0,
                                       vtype=GRB.CONTINUOUS, name='omega8')
    var_dict_SP['omega9'] = SP.addVars(params['Ru_unit_num'], params['time_interval'], lb=0,
                                       vtype=GRB.CONTINUOUS, name='omega9')
    var_dict_SP['omega10'] = SP.addVars(params['Ru_unit_num'], params['time_interval'], lb=0,
                                       vtype=GRB.CONTINUOUS, name='omega10')
    var_dict_SP['omega11'] = SP.addVars(params['Su_unit_num'], params['time_interval'], lb=0,
                                       vtype=GRB.CONTINUOUS, name='omega11')
    var_dict_SP['omega12'] = SP.addVars(params['Ru_unit_num'], params['time_interval'], lb=0,
                                       vtype=GRB.CONTINUOUS, name='omega12')
    var_dict_SP['omega13'] = SP.addVars(params['time_interval'], lb=0,
                                       vtype=GRB.CONTINUOUS, name='omega13')
    var_dict_SP['omega14'] = SP.addVars(params['time_interval'], lb=0,
                                       vtype=GRB.CONTINUOUS, name='omega14')
    var_dict_SP['omega15'] = SP.addVars(params['time_interval'], lb=0,
                                       vtype=GRB.CONTINUOUS, name='omega15')
    var_dict_SP['omega16'] = SP.addVar(lb=-GRB.INFINITY,ub=GRB.INFINITY,vtype=GRB.CONTINUOUS, name='omega16')
    var_dict_SP['omega17'] = SP.addVars(params['time_interval'],lb=-GRB.INFINITY,ub=GRB.INFINITY,
                                       vtype=GRB.CONTINUOUS, name='omega17')
    var_dict_SP['omega18'] = SP.addVars(params['time_interval'], lb=0,
                                        vtype=GRB.CONTINUOUS, name='omega18')
    var_dict_SP['omega19'] = SP.addVars(params['time_interval'], lb=0,
                                        vtype=GRB.CONTINUOUS, name='omega19')
    var_dict_SP['omega20'] = SP.addVars(params['Su_unit_num'], params['time_interval'],lb=0, vtype=GRB.CONTINUOUS,name='omega20')
    var_dict_SP['omega21'] = SP.addVars(params['Ru_unit_num'], params['time_interval'], lb=0, vtype=GRB.CONTINUOUS,
                                        name='omega21')
    var_dict_SP['v1'] = SP.addVars(params['Su_unit_num'], params['time_interval'], vtype=GRB.BINARY,
                                       name='v1')
    var_dict_SP['v2'] = SP.addVars(params['Su_unit_num'], params['time_interval'],
                                       vtype=GRB.BINARY, name='v2')
    var_dict_SP['v3'] = SP.addVars(params['Su_unit_num'], params['La'], params['time_interval'],
                                       vtype=GRB.BINARY, name='v3')
    var_dict_SP['v4'] = SP.addVars(params['Ru_unit_num'], params['time_interval'],
                                       vtype=GRB.BINARY, name='v4')
    var_dict_SP['v5'] = SP.addVars(params['Ru_unit_num'], params['time_interval'],
                                       vtype=GRB.BINARY, name='v5')
    var_dict_SP['v6'] = SP.addVars(params['Ru_unit_num'], params['Lb'], params['time_interval'],
                                       vtype=GRB.BINARY, name='v6')
    var_dict_SP['v7'] = SP.addVars(params['Su_unit_num'], params['time_interval'],
                                       vtype=GRB.BINARY, name='v7')
    var_dict_SP['v8'] = SP.addVars(params['Su_unit_num'], params['time_interval'],
                                       vtype=GRB.BINARY, name='v8')
    var_dict_SP['v9'] = SP.addVars(params['Ru_unit_num'], params['time_interval'],
                                       vtype=GRB.BINARY, name='v9')
    var_dict_SP['v10'] = SP.addVars(params['Ru_unit_num'], params['time_interval'],
                                        vtype=GRB.BINARY, name='v10')
    var_dict_SP['v11'] = SP.addVars(params['Su_unit_num'], params['time_interval'],
                                        vtype=GRB.BINARY, name='v11')
    var_dict_SP['v12'] = SP.addVars(params['Ru_unit_num'], params['time_interval'],
                                        vtype=GRB.BINARY, name='v12')
    var_dict_SP['v13'] = SP.addVars(params['time_interval'],
                                        vtype=GRB.BINARY, name='v13')
    var_dict_SP['v14'] = SP.addVars(params['time_interval'],
                                        vtype=GRB.BINARY, name='v14')
    var_dict_SP['v15'] = SP.addVars(params['time_interval'],
                                        vtype=GRB.BINARY, name='v15')
    var_dict_SP['v18'] = SP.addVars(params['time_interval'],
                                    vtype=GRB.BINARY, name='v18')
    var_dict_SP['v19'] = SP.addVars(params['time_interval'],
                                    vtype=GRB.BINARY, name='v19')
    var_dict_SP['v16'] = SP.addVars(params['Su_unit_num'], params['time_interval'],
                                        vtype=GRB.BINARY, name='v16')
    var_dict_SP['v17'] = SP.addVars(params['Ru_unit_num'], params['time_interval'],
                                    vtype=GRB.BINARY, name='v17')
    var_dict_SP['v21'] = SP.addVars(params['Su_unit_num'], params['time_interval'], vtype=GRB.BINARY,
                                   name='v21')
    var_dict_SP['v22'] = SP.addVars(params['Su_unit_num'], params['time_interval'],
                                   vtype=GRB.BINARY, name='v22')
    var_dict_SP['v23'] = SP.addVars(params['Su_unit_num'], params['La'], params['time_interval'],
                                   vtype=GRB.BINARY, name='v23')
    var_dict_SP['v24'] = SP.addVars(params['Ru_unit_num'], params['time_interval'],
                                   vtype=GRB.BINARY, name='v24')
    var_dict_SP['v25'] = SP.addVars(params['Ru_unit_num'], params['time_interval'],
                                   vtype=GRB.BINARY, name='v25')
    var_dict_SP['v26'] = SP.addVars(params['Ru_unit_num'], params['Lb'], params['time_interval'],
                                   vtype=GRB.BINARY, name='v26')
    var_dict_SP['v27'] = SP.addVars(params['Su_unit_num'], params['time_interval'],
                                   vtype=GRB.BINARY, name='v27')
    var_dict_SP['v28'] = SP.addVars(params['Ru_unit_num'], params['time_interval'],
                                   vtype=GRB.BINARY, name='v28')
    var_dict_SP['v29'] = SP.addVars(params['time_interval'],
                                    vtype=GRB.BINARY, name='v29')
    SP.update()



    return var_dict_SP

def add_SP_constr(var,Xa,Xb,Vdc):
    Ca = var['Ca']
    Cb = var['Cb']
    PA = var['PA']
    PB = var['PB']
    ca = var['ca']
    cb = var['cb']
    Cw = var['Cw']
    cw = var['cw']
    Pwcur = var['Pwcur']
    deltaa = var['deltaa']
    deltab = var['deltab']
    PAMAX = var['PAMAX']
    PBMAX = var['PBMAX']
    Pdc = var['Pdc']
    Pwact = var['Pwact']
    u = var['u']
    Dt = cases.Dt
    sigma = cases.sigma
    Rdc = line.Rdc
    Pdcmin = line.Pdcmin
    Pdcmax = line.Pdcmax
    Pwpred = sne.Pwpred
    Q = line.Q
    delta = line.delta

    Pamax = su.Pmax
    Pamin = su.Pmin
    Pbmax = ru.Pmax
    Pbmin = ru.Pmin
    suRU = su.RU
    ruRU = ru.RU


    T = params['T']
    La = params['La']
    Lb = params['Lb']
    J = params['su_unit_num']
    K = params['ru_unit_num']

    suc1 = su.c1
    suc2 = su.c2
    suc3 = su.c3
    sutv1 = su.tv1
    sutv2 = su.tv2
    ruc1 = ru.c1
    ruc2 = ru.c2
    ruc3 = ru.c3
    rutv1 = ru.tv1
    rutv2 = ru.tv2

    suNL = len(La)
    ruNL = len(Lb)

    FA = np.column_stack([suc1, suc2, suc3])
    FB = np.column_stack([ruc1, ruc2, ruc3])
    TVA = np.column_stack([sutv1, sutv2])
    TVB = np.column_stack([rutv1, rutv2])


    SP.addConstrs(deltaa[j, 0, t] <= TVA[j, 0] for t in T for j in J) #
    SP.addConstrs(deltaa[j, suNL - 1, t] <= Pamax[j] - TVA[j, suNL - 2] for t in T for j in J) #
    SP.addConstrs(deltaa[j, l, t] <= TVA[j, l] - TVA[j, l - 1] for j in J for t in T for l in La[1:suNL - 1]) #





    SP.addConstrs(deltab[k, 0, t] <= TVB[k, 0] for t in T for k in K) #
    SP.addConstrs(deltab[k, ruNL - 1, t] <= Pbmax[k] - TVB[k, ruNL - 2] for t in T for k in K) #
    SP.addConstrs(deltab[k, l, t] <= TVB[k, l] - TVB[k, l - 1] for k in K for t in T for l in Lb[1:suNL - 1]) #



    SP.addConstrs(gp.quicksum(deltaa[j,l,t] for l in La) <= PAMAX[j, t] for t in T for j in J) #
    SP.addConstrs(gp.quicksum(deltaa[j,l,t] for l in La) >= Pamin[j] * Xa[j, t] for t in T for j in J) #
    SP.addConstrs(gp.quicksum(deltab[k,l,t] for l in Lb) <= PBMAX[k, t] for t in T for k in K) #
    SP.addConstrs(gp.quicksum(deltab[k,l,t] for l in Lb) >= Pbmin[k] * Xb[k, t] for k in K for t in T) #
    SP.addConstrs(PAMAX[j, t] <= Pamax[j] * Xa[j, t] for j in J for t in T) #
    SP.addConstrs(PBMAX[k, t] <= Pbmax[k] * Xb[k, t] for k in K for t in T) #
    SP.addConstrs(
        PAMAX[j, t] <= gp.quicksum(deltaa[j,l,t-1] for l in La) + suRU[j] * Xa[j, t - 1] + Pamin[j] * (Xa[j, t] - Xa[j, t - 1]) + Pamax[j] * (
                1 - Xa[j, t]) for j in J for t in T[1:]) #
    SP.addConstrs(
        PBMAX[k, t] <= gp.quicksum(deltab[k,l,t-1] for l in Lb) + ruRU[k] * Xb[k, t - 1] + Pbmin[k] * (Xb[k, t] - Xb[k, t - 1]) + Pbmax[k] * (
                1 - Xb[k, t]) for k in K for t in T[1:]) #
    SP.addConstrs(gp.quicksum(deltaa[j,l,t] for l in La for j in J) + Pwpred[t] - Pwcur[t] - (gp.quicksum(deltaa[j,l,t-1] for l in La for j in J) + Pwpred[t-1] - Pwcur[t-1]) >= -Vdc[t] * Rdc for t in T[1:])
    SP.addConstrs(gp.quicksum(deltaa[j,l,t] for l in La for j in J) + Pwpred[t] - Pwcur[t] - (gp.quicksum(deltaa[j,l,t-1] for l in La for j in J) + Pwpred[t-1] - Pwcur[t-1]) <= Vdc[t] * Rdc for t in T[1:])
    SP.addConstrs(gp.quicksum(deltaa[j,l,t] for l in La for j in J) + Pwpred[t] - Pwcur[t] >= Pdcmin for t in T)
    SP.addConstrs(gp.quicksum(deltaa[j,l,t] for l in La for j in J) + Pwpred[t] - Pwcur[t] <= Pdcmax for t in T)
    SP.addConstr(gp.quicksum(gp.quicksum(deltaa[j,l,t] for l in La for j in J) + Pwpred[t] - Pwcur[t] * delta for t in T) == Q)
    SP.addConstrs(Pwpred[t] - Pwcur[t] >= 0 for t in T)


    SP.addConstrs(gp.quicksum(deltab[k,l,t] for l in Lb for k in K) + gp.quicksum(deltaa[j,l,t] for l in La for j in J) + Pwpred[t] - Pwcur[t] == Dt[t] +u[t] * sigma[t] for t in T)

def add_SP_Duality_constr(var):
    omega1 = var['omega1']
    omega2 = var['omega2']
    omega3 = var['omega3']
    omega4 = var['omega4']
    omega5 = var['omega5']
    omega6 = var['omega6']
    omega7 = var['omega7']
    omega8 = var['omega8']
    omega9 = var['omega9']
    omega10 = var['omega10']
    omega11 = var['omega11']
    omega12 = var['omega12']
    omega13 = var['omega13']
    omega14 = var['omega14']
    omega15 = var['omega15']
    omega16 = var['omega16']
    omega17 = var['omega17']
    omega18 = var['omega18']
    omega19 = var['omega19']
    omega20 = var['omega20']
    omega21 = var['omega21']

    suc1 = su.c1
    suc2 = su.c2
    suc3 = su.c3
    ruc1 = ru.c1
    ruc2 = ru.c2
    ruc3 = ru.c3

    T = params['T']
    La = params['La']
    Lb = params['Lb']
    J = params['su_unit_num']
    K = params['ru_unit_num']
    suNL = len(La)
    ruNL = len(Lb)

    FA = np.column_stack([suc1, suc2, suc3])
    FB = np.column_stack([ruc1, ruc2, ruc3])

    SP.addConstrs(-omega1[j, t] - omega7[j, t] + omega8[j, t] + omega11[j, t + 1] + omega13[t]
                  - omega13[t + 1] - omega14[t] + omega14[t + 1] + omega16 + omega17[t] + omega18[t] - omega19[t] <= FA[
                      j, 0] for j in J for t in T[1:23])
    SP.addConstrs(-omega1[j, 0] - omega7[j, 0] + omega8[j, 0] + omega11[j, 1]
                  - omega13[1] + omega14[1] + omega16 + omega17[0] + omega18[0] - omega19[0] <= FA[j, 0] for j in J)
    SP.addConstrs(-omega1[j, 23] - omega7[j, 23] + omega8[j, 23] + omega13[23]
                  - omega14[23] + omega16 + omega17[23] + omega18[23] - omega19[23] <= FA[j, 0] for j in J)
    SP.addConstrs(-omega2[j, t] - omega7[j, t] + omega8[j, t] + omega11[j, t + 1] + omega13[t]
                  - omega13[t + 1] - omega14[t] + omega14[t + 1] + omega16 + omega17[t] + omega18[t] - omega19[t] <= FA[
                      j, suNL - 1] for j in J for t
                  in T[1:23])
    SP.addConstrs(-omega2[j, 0] - omega7[j, 0] + omega8[j, 0] + omega11[j, 1]
                  - omega13[1] + omega14[1] + omega16 + omega17[0] + omega18[0] - omega19[0] <= FA[j, suNL - 1] for j in
                  J)
    SP.addConstrs(-omega2[j, 23] - omega7[j, 23] + omega8[j, 23] + omega13[23]
                  - omega14[23] + omega16 + omega17[23] + omega18[23] - omega19[23] <= FA[j, suNL - 1] for j in J)
    SP.addConstrs(-omega3[j, l, t] - omega7[j, t] + omega8[j, t] + omega11[j, t + 1] + omega13[t]
                  - omega13[t + 1] - omega14[t] + omega14[t + 1] + omega16 + omega17[t] + omega18[t] - omega19[t] <= FA[
                      j, l] for j in J
                  for t in T[1:23] for l in La[1:suNL - 1])
    SP.addConstrs(-omega3[j, l, 0] - omega7[j, 0] + omega8[j, 0] + omega11[j, 1]
                  - omega13[1] + omega14[1] + omega16 + omega17[0] + omega18[0] - omega19[0] <= FA[j, l] for j in J for
                  l in La[1:suNL - 1])
    SP.addConstrs(-omega3[j, l, 23] - omega7[j, 23] + omega8[j, 23] + omega13[23]
                  - omega14[23] + omega16 + omega17[23] + omega18[23] - omega19[23] <= FA[j, l] for j in J for l in
                  La[1:suNL - 1])

    SP.addConstrs(
        -omega4[k, t] - omega9[k, t] + omega10[k, t] + omega12[k, t + 1] + omega17[t] <= FB[k, 0] for k in K for t in
        T[:23])
    SP.addConstrs(
        -omega4[k, 23] - omega9[k, 23] + omega10[k, 23] + omega17[23] <= FB[k, 0] for k in K)
    SP.addConstrs(
        -omega5[k, t] - omega9[k, t] + omega10[k, t] + omega12[k, t + 1] + omega17[t] <= FB[k, ruNL - 1] for k in K for
        t in
        T[:23])
    SP.addConstrs(
        -omega5[k, 23] - omega9[k, 23] + omega10[k, 23] + omega17[23] <= FB[k, ruNL - 1] for k in K)
    SP.addConstrs(
        -omega6[k, l, t] - omega9[k, t] + omega10[k, t] + omega12[k, t + 1] + omega17[t] <= FB[k, l] for k in K for t in
        T[:23] for l in Lb[1:ruNL - 1])
    SP.addConstrs(
        -omega6[k, l, 23] - omega9[k, 23] + omega10[k, 23] + omega17[23] <= FB[k, l] for k in K for l in Lb[1:ruNL - 1])

    SP.addConstrs(omega7[j, t] - omega11[j, t] - omega20[j, t] <= 0 for j in J for t in T[1:])
    SP.addConstrs(omega7[j, 0] - omega20[j, 0] <= 0 for j in J)

    SP.addConstrs(omega9[k, t] - omega12[k, t] - omega21[k, t] <= 0 for k in K for t in T[1:])
    SP.addConstrs(omega9[k, 0] - omega21[k, 0] <= 0 for k in K)

    SP.addConstrs(
        -omega13[t] + omega13[t + 1] + omega14[t] - omega14[t + 1] - omega16 - omega17[t] - omega18[t] +
        omega19[t] -omega15[t] <= line.cw for t in T[1:23])
    SP.addConstr(omega13[1] - omega14[1] - omega16 - omega17[0] - omega18[0] + omega19[0] -omega15[0] <= line.cw)
    SP.addConstr(
        -omega13[23] + omega14[23] - omega16 - omega17[23] - omega18[23] + omega19[23] -omega15[23] <= line.cw)

def add_SP_KKT1_constr(var,Xa,Xb,Vdc):

    deltaa = var['deltaa']
    deltab = var['deltab']
    PAMAX = var['PAMAX']
    PBMAX = var['PBMAX']

    Pwcur = var['Pwcur']
    omega1 = var['omega1']
    omega2 = var['omega2']
    omega3 = var['omega3']
    omega4 = var['omega4']
    omega5 = var['omega5']
    omega6 = var['omega6']
    omega7 = var['omega7']
    omega8 = var['omega8']
    omega9 = var['omega9']
    omega10 = var['omega10']
    omega11 = var['omega11']
    omega12 = var['omega12']
    omega13 = var['omega13']
    omega14 = var['omega14']
    omega15 = var['omega15']
    omega18 = var['omega18']
    omega19 = var['omega19']
    omega20 = var['omega20']
    omega21 = var['omega21']

    v1 = var['v1']
    v2 = var['v2']
    v3 = var['v3']
    v4 = var['v4']
    v5 = var['v5']
    v6 = var['v6']
    v7 = var['v7']
    v8 = var['v8']
    v9 = var['v9']
    v10 = var['v10']
    v11 = var['v11']
    v12 = var['v12']
    v13 = var['v13']
    v14 = var['v14']
    v15 = var['v15']
    v18 = var['v18']
    v19 = var['v19']
    v16 = var['v16']
    v17 = var['v17']


    Rdc = line.Rdc
    Pdcmin = line.Pdcmin
    Pdcmax = line.Pdcmax
    Pwpred = sne.Pwpred


    Pamax = su.Pmax
    Pamin = su.Pmin
    Pbmax = ru.Pmax
    Pbmin = ru.Pmin
    suRU = su.RU
    ruRU = ru.RU

    T = params['T']
    La = params['La']
    Lb = params['Lb']
    J = params['su_unit_num']
    K = params['ru_unit_num']


    sutv1 = su.tv1
    sutv2 = su.tv2

    rutv1 = ru.tv1
    rutv2 = ru.tv2


    suNL = len(La)
    ruNL = len(Lb)


    TVA = np.column_stack([sutv1, sutv2])
    TVB = np.column_stack([rutv1, rutv2])

    SP.addConstrs(-deltaa[j, 0, t] + TVA[j, 0] <= M * (1 - v1[j, t]) for j in J for t in T)
    SP.addConstrs(omega1[j, t] <= M * v1[j, t] for j in J for t in T)
    SP.addConstrs(-deltaa[j, suNL - 1, t] + Pamax[j] - TVA[j, suNL - 2] <= M * (1 - v2[j, t]) for j in J for t in T)
    SP.addConstrs(omega2[j, t] <= M * v2[j, t] for j in J for t in T)
    SP.addConstrs(
        -deltaa[j, l, t] + TVA[j, l] - TVA[j, l - 1] <= M * (1 - v3[j, l, t]) for j in J for l in La[1:suNL - 1] for t
        in T)
    SP.addConstrs(omega3[j, l, t] <= M * v3[j, l, t] for j in J for l in La[1:suNL - 1] for t in T)
    SP.addConstrs(-deltab[k, 0, t] + TVB[k, 0] <= M * (1 - v4[k, t]) for k in K for t in T)
    SP.addConstrs(omega4[k, t] <= M * v4[k, t] for k in K for t in T)
    SP.addConstrs(-deltab[k, ruNL - 1, t] + Pbmax[k] - TVB[k, ruNL - 2] <= M * (1 - v5[k, t]) for k in K for t in T)
    SP.addConstrs(omega5[k, t] <= M * v5[k, t] for k in K for t in T)
    SP.addConstrs(-deltab[k, l, t] + TVB[k, l] - TVB[k, l - 1] <= M * (1 - v6[k, l, t]) for t in T for k in K for l in
                  Lb[1:ruNL - 1])
    SP.addConstrs(omega6[k, l, t] <= M * v6[k, l, t] for k in K for l in Lb[1:ruNL - 1] for t in T)
    SP.addConstrs(-gp.quicksum(deltaa[j, l, t] for l in La) + PAMAX[j, t] <= M * (1 - v7[j, t]) for t in T for j in J)
    SP.addConstrs(omega7[j, t] <= M * v7[j, t] for j in J for t in T)
    SP.addConstrs(
        gp.quicksum(deltaa[j, l, t] for l in La) - Pamin[j] * Xa[j, t] <= M * (1 - v8[j, t]) for j in J for t in T)
    SP.addConstrs(omega8[j, t] <= M * v8[j, t] for j in J for t in T)
    SP.addConstrs(-gp.quicksum(deltab[k, l, t] for l in Lb) + PBMAX[k, t] <= M * (1 - v9[k, t]) for k in K for t in T)
    SP.addConstrs(omega9[k, t] <= M * v9[k, t] for k in K for t in T)
    SP.addConstrs(
        gp.quicksum(deltab[k, l, t] for l in Lb) - Pbmin[k] * Xb[k, t] <= M * (1 - v10[k, t]) for k in K for t in T)
    SP.addConstrs(omega10[k, t] <= M * v10[k, t] for k in K for t in T)
    SP.addConstrs(gp.quicksum(deltaa[j, l, t-1] for l in La) - PAMAX[j, t] + suRU[j] * Xa[j, t - 1] + Pamin[j] * (
            Xa[j, t] - Xa[j, t - 1]) + Pamax[j] * (
                          1 - Xa[j, t]) <= M * (1 - v11[j, t]) for j in J for t in T[1:])
    SP.addConstrs(omega11[j, t] <= M * v11[j, t] for j in J for t in T[1:])
    SP.addConstrs(gp.quicksum(deltab[k, l, t-1] for l in Lb) - PBMAX[k, t] + ruRU[k] * Xb[k, t - 1] + Pbmin[k] * (
            Xb[k, t] - Xb[k, t - 1]) + Pbmax[k] * (
                          1 - Xb[k, t]) <= M * (1 - v12[k, t]) for k in K for t in T[1:])
    SP.addConstrs(omega12[k, t] <= M * v12[k, t] for k in K for t in T[1:])
    SP.addConstrs(gp.quicksum(deltaa[j, l, t] - deltaa[j, l, t - 1] for j in J for l in La) -Pwcur[t] + Pwcur[t-1] +Vdc[t] * Rdc + Pwpred[t] - Pwpred[t-1] <= M * (1 - v13[t]) for t in T[1:])
    SP.addConstrs(omega13[t] <= M * v13[t] for t in T[1:])
    SP.addConstrs(
        gp.quicksum(-deltaa[j, l, t] + deltaa[j, l, t - 1] for j in J for l in La) +Pwcur[t] - Pwcur[t-1] + Vdc[t] * Rdc - Pwpred[t] + Pwpred[t-1] <= M * (
                1 - v14[t]) for t in T[1:])
    SP.addConstrs(omega14[t] <= M * v14[t] for t in T[1:])

    SP.addConstrs(Pwpred[t] - Pwcur[t] <= M * (1-v15[t]) for t in T)
    SP.addConstrs(omega15[t] <= M * v15[t] for t in T)

    SP.addConstrs(
        gp.quicksum(deltaa[j, l, t] for j in J for l in La) -Pwcur[t] - Pdcmin + Pwpred[t] <= M * (1 - v18[t]) for t in T)
    SP.addConstrs(omega18[t] <= M * v18[t] for t in T)
    SP.addConstrs(
        gp.quicksum(-deltaa[j, l, t] for j in J for l in La) + Pwcur[t] - Pwpred[t] + Pdcmax <= M * (1 - v19[t]) for t in T)
    SP.addConstrs(omega19[t] <= M * v19[t] for t in T)
    SP.addConstrs(-PAMAX[j,t] + Pamax[j] * Xa[j,t] <= M * (1-v16[j,t]) for t in T for j in J)
    SP.addConstrs(omega20[j,t] <= M * v16[j,t] for t in T for j in J)
    SP.addConstrs(-PBMAX[k,t] + Pbmax[k] * Xb[k,t] <= M * (1-v17[k,t]) for k in K for t in T)
    SP.addConstrs(omega21[k,t] <= M * v17[k,t] for k in K for t in T)

def add_SP_KKT2_constr(var):
    omega1 = var['omega1']
    omega2 = var['omega2']
    omega3 = var['omega3']
    omega4 = var['omega4']
    omega5 = var['omega5']
    omega6 = var['omega6']
    omega7 = var['omega7']
    omega8 = var['omega8']
    omega9 = var['omega9']
    omega10 = var['omega10']
    omega11 = var['omega11']
    omega12 = var['omega12']
    omega13 = var['omega13']
    omega14 = var['omega14']
    omega15 = var['omega15']
    omega16 = var['omega16']
    omega17 = var['omega17']
    omega18 = var['omega18']
    omega19 = var['omega19']
    omega20 = var['omega20']
    omega21 = var['omega21']
    deltaa = var['deltaa']
    deltab = var['deltab']
    PAMAX = var['PAMAX']
    PBMAX = var['PBMAX']
    Pwact = var['Pwact']
    Pwcur = var['Pwcur']
    v21 = var['v21']
    v22 = var['v22']
    v23 = var['v23']
    v24 = var['v24']
    v25 = var['v25']
    v26 = var['v26']
    v27 = var['v27']
    v28 = var['v28']
    v29 = var['v29']

    suc1 = su.c1
    suc2 = su.c2
    suc3 = su.c3
    ruc1 = ru.c1
    ruc2 = ru.c2
    ruc3 = ru.c3

    T = params['T']
    La = params['La']
    Lb = params['Lb']
    J = params['su_unit_num']
    K = params['ru_unit_num']
    suNL = len(La)
    ruNL = len(Lb)

    FA = np.column_stack([suc1, suc2, suc3])
    FB = np.column_stack([ruc1, ruc2, ruc3])

    SP.addConstrs(omega1[j, t] + omega7[j, t] - omega8[j, t] - omega11[j, t+1] - omega13[t]
                  + omega13[t + 1] + omega14[t] - omega14[t + 1] - omega16 - omega17[t] - omega18[t] + omega19[t] + FA[
                      j, 0] <= M * (1 - v21[j, t]) for j in J for t in T[1:23])
    SP.addConstrs(omega1[j, 0] + omega7[j, 0] - omega8[j, 0] - omega11[j,1]
                  + omega13[1] - omega14[1] - omega16 - omega17[0] - omega18[0] + omega19[0] + FA[j, 0] <= M * (
                              1 - v21[j, 0]) for j in J)
    SP.addConstrs(omega1[j, 23] + omega7[j, 23] - omega8[j, 23] - omega13[23]
                  + omega14[23] - omega16 - omega17[23] - omega18[23] + omega19[23] + FA[j, 0] <= M * (1 - v21[j, 23])
                  for j in J)
    SP.addConstrs(deltaa[j, 0, t] <= M * v21[j, t] for j in J for t in T)

    SP.addConstrs(omega2[j, t] + omega7[j, t] - omega8[j, t] - omega11[j, t+1] - omega13[t]
                  + omega13[t + 1] + omega14[t] - omega14[t + 1] - omega16 - omega17[t] - omega18[t] + omega19[t] + FA[
                      j, suNL - 1] <= M * (1 - v22[j, t]) for j in J for t
                  in T[1:23])
    SP.addConstrs(omega2[j, 0] + omega7[j, 0] - omega8[j, 0] - omega11[j,1]
                  + omega13[1] - omega14[1] - omega16 - omega17[0] - omega18[0] + omega19[0] + FA[j, suNL - 1] <= M * (
                              1 - v22[j, 0]) for j in
                  J)
    SP.addConstrs(omega2[j, 23] + omega7[j, 23] - omega8[j, 23] - omega13[23]
                  + omega14[23] - omega16 - omega17[23] - omega18[23] + omega19[23] + FA[j, suNL - 1] <= M * (
                              1 - v22[j, 23]) for j in J)
    SP.addConstrs(deltaa[j, suNL - 1, t] <= M * v22[j, t] for j in J for t in T)

    SP.addConstrs(omega3[j, l, t] + omega7[j, t] - omega8[j, t] - omega11[j, t+1] - omega13[t]
                  + omega13[t + 1] + omega14[t] - omega14[t + 1] - omega16 - omega17[t] - omega18[t] + omega19[t] + FA[
                      j, l] <= M * (1 - v23[j, l, t]) for j in J
                  for t in T[1:23] for l in La[1:suNL - 1])
    SP.addConstrs(omega3[j, l, 0] + omega7[j, 0] - omega8[j, 0] -omega11[j,1]
                  + omega13[1] - omega14[1] - omega16 - omega17[0] - omega18[0] + omega19[0] + FA[j, l] <= M * (
                              1 - v23[j, l, 0]) for j in J for
                  l in La[1:suNL - 1])
    SP.addConstrs(omega3[j, l, 23] + omega7[j, 23] - omega8[j, 23] - omega13[23]
                  + omega14[23] - omega16 - omega17[23] - omega18[23] + omega19[23] + FA[j, l] <= M * (
                              1 - v23[j, l, 23]) for j in J for l in
                  La[1:suNL - 1])
    SP.addConstrs(deltaa[j, l, t] <= M * v23[j, l, t] for j in J for l in La[1:suNL - 1] for t in T)

    SP.addConstrs(
        omega4[k, t] + omega9[k, t] - omega10[k, t] - omega12[k, t+1] - omega17[t] + FB[k, 0] <= M * (1 - v24[k, t]) for k
        in K for t in T[:23])
    SP.addConstrs(
        omega4[k, 23] + omega9[k, 23] - omega10[k, 23] - omega17[23] + FB[k, 0] <= M * (1 - v24[k, 23]) for k in K)
    SP.addConstrs(deltab[k, 0, t] <= M * v24[k, t] for k in K for t in T)

    SP.addConstrs(
        omega5[k, t] + omega9[k, t] - omega10[k, t] - omega12[k, t+1] - omega17[t] + FB[k, ruNL - 1] <= M * (
                    1 - v25[k, t]) for k in K for t in
        T[:23])
    SP.addConstrs(
        omega5[k, 23] + omega9[k, 23] - omega10[k, 23] - omega17[23] + FB[k, ruNL - 1] <= M * (1 - v25[k, 23]) for k in K)
    SP.addConstrs(deltab[k, ruNL - 1, t] <= M * v25[k, t] for k in K for t in T)

    SP.addConstrs(
        omega6[k, l, t] + omega9[k, t] - omega10[k, t] - omega12[k, t+1] - omega17[t] + FB[k, l] <= M * (1 - v26[k, l, t])
        for k in K for t in
        T[:23] for l in Lb[1:ruNL - 1])
    SP.addConstrs(
        omega6[k, l, 23] + omega9[k, 23] - omega10[k, 23] - omega17[23] + FB[k, l] <= M * (1 - v26[k, l, 23]) for k in K for
        l in Lb[1:ruNL - 1])
    SP.addConstrs(deltab[k, l, t] <= M * v26[k, l, t] for k in K for l in Lb[1:ruNL - 1] for t in T)

    SP.addConstrs(-omega7[j, t] + omega11[j, t] + omega20[j,t] <= M * (1 - v27[j, t]) for j in J for t in T[1:])
    SP.addConstrs(-omega7[j, 0] +omega20[j,0] <= M * (1 - v27[j, 0]) for j in J)
    SP.addConstrs(PAMAX[j, t] <= M * v27[j, t] for j in J for t in T)

    SP.addConstrs(-omega9[k, t] + omega12[k, t] +omega21[k,t] <= M * (1 - v28[k, t]) for k in K for t in T[1:])
    SP.addConstrs(-omega9[k, 0] +omega21[k,0] <= M * (1 - v28[k, 0]) for k in K)
    SP.addConstrs(PBMAX[k, t] <= M * v28[k, t] for k in K for t in T)

    SP.addConstrs(
        omega13[t] - omega13[t + 1] - omega14[t] + omega14[t + 1] + omega16 + omega17[t] + omega18[t] -
        omega19[t] +omega15[t] + line.cw <= M * (1 - v29[t]) for t in T[1:23])
    SP.addConstr(
        -omega13[1] + omega14[1] + omega16 + omega17[0] + omega18[0] - omega19[0] +omega15[0]+ line.cw <= M * (
                    1 - v29[0]))
    SP.addConstr(
        omega13[23] - omega14[23] + omega16 + omega17[23] + omega18[23] - omega19[23] +omega15[23]+ line.cw <= M * (
                    1 - v29[23]))
    SP.addConstrs(Pwcur[t] <= M * v29[t] for t in T)

def add_SP_constrs(var,Xa,Xb,Vdc):


    add_SP_Duality_constr(var)





def SP_solve(Xa,Xb,Vdc):
    SP.remove(SP.getConstrs())
    var_dict_SP = add_SP_var()
    omega1 = var_dict_SP['omega1']
    omega2 = var_dict_SP['omega2']
    omega3 = var_dict_SP['omega3']
    omega4 = var_dict_SP['omega4']
    omega5 = var_dict_SP['omega5']
    omega6 = var_dict_SP['omega6']
    omega7 = var_dict_SP['omega7']
    omega8 = var_dict_SP['omega8']
    omega9 = var_dict_SP['omega9']
    omega10 = var_dict_SP['omega10']
    omega11 = var_dict_SP['omega11']
    omega12 = var_dict_SP['omega12']
    omega13 = var_dict_SP['omega13']
    omega14 = var_dict_SP['omega14']
    omega15 = var_dict_SP['omega15']
    omega16 = var_dict_SP['omega16']
    omega17 = var_dict_SP['omega17']
    omega18 = var_dict_SP['omega18']
    omega19 = var_dict_SP['omega19']
    omega20 = var_dict_SP['omega20']
    omega21 = var_dict_SP['omega21']
    u = var_dict_SP['u']
    Dt = cases.Dt
    sigma = cases.sigma
    Rdc = line.Rdc
    Pdcmin = line.Pdcmin
    Pdcmax = line.Pdcmax
    Pwpred = sne.Pwpred
    Q = line.Q
    delta = line.delta

    Pamax = su.Pmax
    Pamin = su.Pmin
    Pbmax = ru.Pmax
    Pbmin = ru.Pmin
    suRU = su.RU
    ruRU = ru.RU

    T = params['T']
    La = params['La']
    Lb = params['Lb']
    J = params['su_unit_num']
    K = params['ru_unit_num']

    suc1 = su.c1
    suc2 = su.c2
    suc3 = su.c3
    sutv1 = su.tv1
    sutv2 = su.tv2
    ruc1 = ru.c1
    ruc2 = ru.c2
    ruc3 = ru.c3
    rutv1 = ru.tv1
    rutv2 = ru.tv2

    suNL = len(La)
    ruNL = len(Lb)

    FA = np.column_stack([suc1, suc2, suc3])
    FB = np.column_stack([ruc1, ruc2, ruc3])
    TVA = np.column_stack([sutv1, sutv2])
    TVB = np.column_stack([rutv1, rutv2])
    Pwpred = sne.Pwpred

    SP.setObjective(gp.quicksum(-TVA[j, 0] * omega1[j, t] for j in J for t in T)
                    + gp.quicksum((TVA[j, suNL - 2] - Pamax[j]) * omega2[j, t] for j in J for t in T)
                    + gp.quicksum(
        (TVA[j, l - 1] - TVA[j, l]) * omega3[j, l, t] for j in J for l in La[1:suNL - 1] for t in T)
                    + gp.quicksum(-TVB[k, 0] * omega4[k, t] for k in K for t in T)
                    + gp.quicksum((TVB[k, ruNL - 2] - Pbmax[k]) * omega5[k, t] for k in K for t in T)
                    + gp.quicksum(
        (TVB[k, l - 1] - TVB[k, l]) * omega6[k, l, t] for k in K for l in Lb[1:ruNL - 1] for t in T)
                    + gp.quicksum(Pamin[j] * Xa[j, t] * omega8[j, t] for j in J for t in T)
                    + gp.quicksum(Pbmin[k] * Xb[k, t] * omega10[k, t] for k in K for t in T)
                    + gp.quicksum(
        -(suRU[j] * Xa[j, t - 1] + Pamin[j] * (Xa[j, t] - Xa[j, t - 1]) + Pamax[j] * (1 - Xa[j, t])) * omega11[j, t] for
        j in J for t in T[1:])
                    + gp.quicksum(
        -(ruRU[k] * Xb[k, t - 1] + Pbmin[k] * (Xb[k, t] - Xb[k, t - 1]) + Pbmax[k] * (1 - Xb[k, t])) * omega12[k, t] for
        k in K for t in T[1:])
                    + gp.quicksum((-Vdc[t] * Rdc - Pwpred[t] + Pwpred[t - 1]) * omega13[t] for t in T[1:])
                    + gp.quicksum((-Vdc[t] * Rdc + Pwpred[t] - Pwpred[t - 1]) * omega14[t] for t in T[1:])
                    + gp.quicksum(-Pwpred[t] * omega15[t] for t in T)
                    + (Q - gp.quicksum(Pwpred[t] for t in T)) * omega16
                    + gp.quicksum((Dt[t] - Pwpred[t] + u[t] * sigma[t]) * omega17[t] for t in T)
                    + gp.quicksum((Pdcmin - Pwpred[t]) * omega18[t] for t in T)
                    + gp.quicksum((-Pdcmax + Pwpred[t]) * omega19[t] for t in T)
                    + gp.quicksum(-Pamax[j] * Xa[j, t] * omega20[j, t] for j in J for t in T)
                    + gp.quicksum(-Pbmax[k] * Xb[k, t] * omega21[k, t] for k in K for t in T), gurobipy.GRB.MAXIMIZE)
    add_SP_constrs(var_dict_SP,Xa,Xb,Vdc)



    SP.update()



    SP.optimize()








def data_analysis(crim):
    T = params['T']
    X = range(1,params['time_interval']+1)
    for c in range(crim+1):
        x = X
        y = [u[c,t] for t in T]
        plt.plot(x, y, label=f"变量{c + 1}")
    plt.title("u")
    plt.xlabel("时间(h)")
    plt.ylabel("u")
    plt.legend()
    plt.show()



if __name__ == "__main__":
    UB1 = GRB.INFINITY
    LB1 = -GRB.INFINITY


    T = params['T']
    J = params['su_unit_num']
    K = params['ru_unit_num']
    La = params['La']
    Lb = params['Lb']

    var_dict_SP = add_SP_var()
    crim = 0
    MAX = 20
    L1 = []
    U1 = []
    Num = []
    etas=[]
    spobjs=[]
    suop = su.Cop
    susd = su.Csd
    ruop = ru.Cop
    rusd = ru.Csd
    suV0 = su.V0
    ruV0 = ru.V0
    suUT = su.UT
    ruUT = ru.UT
    suDT = su.DT
    ruDT = ru.DT
    suU0 = su.U0
    suS0 = su.S0
    ruU0 = ru.U0
    ruS0 = ru.S0
    suV0 = su.V0
    ruV0 = ru.V0
    Tia = np.squeeze(np.full((1, len(J)), params['time_interval']))
    Tib = np.squeeze(np.full((1, len(K)), params['time_interval']))
    GA = np.minimum(Tia, (suUT - suU0) * suV0)
    GB = np.minimum(Tib, (ruUT - ruU0) * ruV0)
    GA[GA < 0] = 0
    GB[GB < 0] = 0
    LA = np.minimum(Tia, (suDT - suS0) * (1 - suV0))
    LB = np.minimum(Tib, (ruDT - ruS0) * (1 - ruV0))
    LA[LA < 0] = 0
    LB[LB < 0] = 0
    TM = line.TM
    X = line.X
    time_interval = params['time_interval']
    MP.addConstrs(cau[j, t] >= suop[j] * (xa[j, t] - xa[j, t - 1]) for j in J for t in T[1:])
    MP.addConstrs(cbu[k, t] >= ruop[k] * (xb[k, t] - xb[k, t - 1]) for k in K for t in T[1:])
    MP.addConstrs(cad[j, t] >= susd[j] * (xa[j, t - 1] - xa[j, t]) for j in J for t in T[1:])
    MP.addConstrs(cbd[k, t] >= rusd[k] * (xb[k, t - 1] - xb[k, t]) for k in K for t in T[1:])
    MP.addConstrs(cau[j, 0] >= suop[j] * (xa[j, 0] - suV0[j]) for j in J)
    MP.addConstrs(cbu[k, 0] >= ruop[k] * (xb[k, 0] - ruV0[k]) for k in K)
    MP.addConstrs(cad[j, 0] >= susd[j] * (suV0[j] - xa[j, 0]) for j in J)
    MP.addConstrs(cbd[k, 0] >= rusd[k] * (ruV0[k] - xb[k, 0]) for k in K)
    MP.addConstrs(gp.quicksum(1 - xa[j, t] for t in range(GA[j])) == 0 for j in J)
    MP.addConstrs(
        gp.quicksum(xa[j, t] for t in range(k, k + suUT[j] - 1)) >= suUT[j] * (xa[j, k] - xa[j, k - 1]) for j in J for k
        in range(GA[j] + 1, Tia[j] - suUT[j] + 1))
    MP.addConstrs(gp.quicksum(xa[j, t] - (xa[j, k] - xa[j, k - 1]) for t in T[k:]) >= 0 for j in J for k in
                  range(Tia[j] - suUT[j] + 2, Tia[j]))
    MP.addConstrs(gp.quicksum(1 - xb[k, t] for t in range(GB[k])) == 0 for k in K)
    MP.addConstrs(
        gp.quicksum(xb[n, t] for t in range(k, k + ruUT[n] - 1)) >= ruUT[n] * (xb[n, k] - xb[n, k - 1]) for n in K for k
        in range(GB[n] + 1, Tib[n] - ruUT[n] + 1))
    MP.addConstrs(gp.quicksum(xb[n, t] - (xb[n, k] - xb[n, k - 1]) for t in T[k:]) >= 0 for n in K for k in
                  range(Tib[n] - ruUT[n] + 2, Tib[n]))
    MP.addConstrs(gp.quicksum(xa[j, t] for t in range(LA[j])) == 0 for j in J)
    MP.addConstrs(
        gp.quicksum(1 - xa[j, t] for t in range(k, k + suDT[j] - 1)) >= suDT[j] * (xa[j, k - 1] - xa[j, k]) for j in J
        for k in range(LA[j] + 1, Tia[j] - suDT[j] + 1))
    MP.addConstrs(gp.quicksum(1 - xa[j, t] - (xa[j, k - 1] - xa[j, k]) for t in T[k:]) >= 0 for j in J for k in
                  range(Tia[j] - suDT[j] + 2, Tia[j]))
    MP.addConstrs(gp.quicksum(xb[k, t] for t in range(LB[k])) == 0 for k in K)
    MP.addConstrs(
        gp.quicksum(1 - xb[n, t] for t in range(k, k + ruDT[n] - 1)) >= ruDT[n] * (xb[n, k - 1] - xb[n, k]) for n in K
        for k in range(LB[n] + 1, Tib[n] - ruDT[n] + 1))
    MP.addConstrs(gp.quicksum(1 - xb[n, t] - (xb[n, k - 1] - xb[n, k]) for t in T[k:]) >= 0 for n in K for k in
                  range(Tib[n] - ruDT[n] + 2, Tib[n]))
    MP.addConstrs(gp.quicksum(vdc[t] for t in T[k:k + TM]) <= 1 for k in T[:time_interval - TM])
    MP.addConstr(gp.quicksum(vdc[t] for t in T) <= X)
    for t in T:
        u[crim,t] = 0
    wb = Workbook()
    ws1 = wb.create_sheet(title='PA')
    ws2 = wb.create_sheet(title='PB')
    ws3 = wb.create_sheet(title='Pwact')
    ws4 = wb.create_sheet(title='Pwcur')
    ws5 = wb.create_sheet(title='Pdc')
    ws6 = wb.create_sheet(title='vdc')
    ws7 = wb.create_sheet(title='xa1')
    ws8 = wb.create_sheet(title='xa2')
    ws9 = wb.create_sheet(title='xa3')
    ws10 = wb.create_sheet(title='xa4')
    ws11 = wb.create_sheet(title='xa5')
    ws12 = wb.create_sheet(title='xb1')
    ws13 = wb.create_sheet(title='xb2')
    ws14 = wb.create_sheet(title='xb3')
    ws15 = wb.create_sheet(title='xb4')
    ws16 = wb.create_sheet(title='xb5')
    ws17 = wb.create_sheet(title='u')
    while(UB1 - LB1>=10e-3 and MAX >0):

        MP_solve(u,crim)


        LB1 = MP.ObjVal
        for j in J:
            for t in T:
                Xa[j, t] = xa[j, t].x
        for k in K:
            for t in T:
                Xb[k, t] = xb[k, t].x

        for t in T:
            Vdc[t] = vdc[t].x

        SP_solve(Xa, Xb, Vdc)





        UB1 = min(UB1, LB1 - eta.x + SP.ObjVal)


        for t in T:
            u[crim+1,t] = var_dict_SP['u'][t].x


        PA = []


        for i in T:
            pa = 0
            for j in J:
                pa += var_dict_MP[crim,'PA'][j,i].x
            PA.append(pa)


        for i, value in enumerate(PA, start=1):
            ws1.cell(row=i, column=crim+1, value=value)



        PB = []


        for i in T:
            pb = 0
            for k in K:
                pb += var_dict_MP[crim, 'PB'][k, i].x
            PB.append(pb)

        for i, value in enumerate(PB, start=1):
            ws2.cell(row=i, column=crim + 1, value=value)

        Pwact = []

        for i in T:
            Pwact.append(var_dict_MP[crim,'Pwact'][i].x)

        for i, value in enumerate(Pwact, start=1):
            ws3.cell(row=i, column=crim + 1, value=value)

        Pwcur = []

        for i in T:
            Pwcur.append(var_dict_MP[crim, 'Pwcur'][i].x)

        for i, value in enumerate(Pwcur, start=1):
            ws4.cell(row=i, column=crim + 1, value=value)

        Pdc = []

        for i in T:
            Pdc.append(var_dict_MP[crim, 'Pdc'][i].x)

        for i, value in enumerate(Pdc, start=1):
            ws5.cell(row=i, column=crim + 1, value=value)

        VDC = []

        for i in T:
            VDC.append(vdc[i].x)

        for i, value in enumerate(VDC, start=1):
            ws6.cell(row=i, column=crim + 1, value=value)

        XA1 = []

        for i in T:
            XA1.append(xa[0,i].x)

        for i, value in enumerate(XA1, start=1):
            ws7.cell(row=i, column=crim + 1, value=value)

        XA2 = []

        for i in T:
            XA2.append(xa[1, i].x)

        for i, value in enumerate(XA2, start=1):
            ws8.cell(row=i, column=crim + 1, value=value)

        XA3 = []

        for i in T:
            XA3.append(xa[2, i].x)

        for i, value in enumerate(XA3, start=1):
            ws9.cell(row=i, column=crim + 1, value=value)

        XA4 = []

        for i in T:
            XA4.append(xa[3, i].x)

        for i, value in enumerate(XA4, start=1):
            ws10.cell(row=i, column=crim + 1, value=value)

        XA5 = []

        for i in T:
            XA5.append(xa[4, i].x)

        for i, value in enumerate(XA5, start=1):
            ws11.cell(row=i, column=crim + 1, value=value)

        XB1 = []

        for i in T:
            XB1.append(xb[0, i].x)

        for i, value in enumerate(XB1, start=1):
            ws12.cell(row=i, column=crim + 1, value=value)

        XB2 = []

        for i in T:
            XB2.append(xb[1, i].x)

        for i, value in enumerate(XB2, start=1):
            ws13.cell(row=i, column=crim + 1, value=value)

        XB3 = []

        for i in T:
            XB3.append(xb[2, i].x)

        for i, value in enumerate(XB3, start=1):
            ws14.cell(row=i, column=crim + 1, value=value)

        XB4 = []

        for i in T:
            XB4.append(xb[3, i].x)

        for i, value in enumerate(XB4, start=1):
            ws15.cell(row=i, column=crim + 1, value=value)

        XB5 = []

        for i in T:
            XB5.append(xb[4, i].x)

        for i, value in enumerate(XB5, start=1):
            ws16.cell(row=i, column=crim + 1, value=value)

        U = []

        for i in T:
            U.append(u[crim,i])

        for i, value in enumerate(U, start=1):
            ws17.cell(row=i, column=crim + 1, value=value)




        crim += 1
        MAX -= 1
        L1.append(LB1)
        U1.append(UB1)
        etas.append(eta.x)
        spobjs.append(SP.ObjVal)
        Num.append(SP.NumConstrs)



wb.save('data.xlsx')
print("下界", L1)
print('上界', U1)
print('eta:', etas)
print('Q(y):',spobjs)
print(Num)
data_analysis(crim)
